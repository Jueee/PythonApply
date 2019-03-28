'''
openpyxl


而2007往后版本的Excel多出来个xlsx文件类型，是为了使Excel能存入超过65535行数据（1048576），所以读写xlsx文件需要另一个库叫openpyxl，支持Python3.x
'''
from openpyxl.reader.excel import load_workbook

wb = load_workbook('dd.xlsx')

sheetnames = wb.sheetnames
print(sheetnames)

ws = wb[sheetnames[0]]
print(ws)
print(ws.max_row)


data_dic = {}       # 建立存储数据的字典

# 把数据存到字典中
for rx in range(2,ws.max_row):
    temp_list = []
    pid = ws.cell(row=rx, column = 1).value
    w1 = ws.cell(row=rx, column = 1).value
    w2 = ws.cell(row=rx, column = 2).value
    w3 = ws.cell(row=rx, column = 3).value
    w4 = ws.cell(row=rx, column = 4).value
    temp_list = [w1, w2, w3, w4]

    data_dic[pid] = temp_list

# 打印字典数据个数
print(data_dic)


